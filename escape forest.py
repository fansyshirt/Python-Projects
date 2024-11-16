import random
import time
try:
    import openpyxl as xl
    from openpyxl.chart import BarChart, Reference
except ModuleNotFoundError:
    print("you do not have openpyxl module installed")
    print("you will not be able to save the game")


def introduction():
    print("you regain consciousness")
    print("...")
    time.sleep(4)
    print("you are in a forest")
    print("...")
    time.sleep(4)
    print("everything seems unfamiliar")
    print("...")
    time.sleep(5)
    print("three witches are heard talking")
    print("...")
    time.sleep(5)
    print('"that fool, he (idk need suggestion)"')
    print("...")
    time.sleep(5)
    print('"he wont be a problem to us anymore"')
    print("...")
    time.sleep(5)
    print('"now that the forest is guarded with our minions"')
    print("...")
    time.sleep(5)
    print('"that mortal will never oppose the Elevated One"')
    print("...")
    time.sleep(5)
    print("*cackled laughing*")
    time.sleep(5)
    print()
    print()
    print()


def save_data():
    try:
        file = "escape_forest_save.xlsx"
        workbook = xl.load_workbook(file)
        page = workbook['Sheet1']

        save = HP
        save_cell = page.cell(2, 2)
        save_cell.value = save

        save = HP_max
        save_cell = page.cell(3, 2)
        save_cell.value = save

        save = attack
        save_cell = page.cell(4, 2)
        save_cell.value = save

        save = defence
        save_cell = page.cell(5, 2)
        save_cell.value = save

        save = regen_pot_count
        save_cell = page.cell(6, 2)
        save_cell.value = save

        save = candy
        save_cell = page.cell(7, 2)
        save_cell.value = save

        save = witch_B_dead
        save_cell = page.cell(8, 2)
        save_cell.value = save

        save = witch_U_dead
        save_cell = page.cell(9, 2)
        save_cell.value = save

        save = witch_G_dead
        save_cell = page.cell(10, 2)
        save_cell.value = save

        save = summon_Elevated
        save_cell = page.cell(11, 2)
        save_cell.value = save

        workbook.save(file)
    except ValueError:
        print("bruh this shouldn't happen")
    except NameError:
        return 1


def summon_elevated_one():
    print("the three tokens the witches dropped snap together")
    print("...")
    time.sleep(4)
    print("it starts to glow and burn your hand")
    print("...")
    time.sleep(4)
    print("you drop it and it sinks into the ground")
    print("...")
    time.sleep(4)
    print("a voice booms,")
    print('"WHO DARES TO SUMMON THE ELEVATED ONE"')
    print("")
    time.sleep(4)
    print('"YOU PUNY MORTAL?"')
    print("")
    time.sleep(6)
    print('"YOU HAVE AWAKENED ME FOR NOTHING"')
    print("")
    time.sleep(4)
    print('"I WILL NOW REMOVE YOU FROM EXISTANCE"')
    print("")
    time.sleep(4)


def credits():
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print("thank you for playing Escape Forest")
    time.sleep(1)
    print()
    time.sleep(1)
    print()
    time.sleep(1)
    print()
    time.sleep(1)
    print("Creators:")
    time.sleep(1)
    print()
    time.sleep(1)
    print("Peter Babych")


rng1 = 0  # find monster
rng2 = 0  # candies found
rng3 = 0  # monster level
rng4 = 0  # attack variance (not implemented)
rng5 = 0  # candies dropped

command = ""

HP = 100
HP_max = 100
attack = 10
defence = 10

regen_pot_count = 0
candy = 0

witch_B_dead = 0
witch_U_dead = 0
witch_G_dead = 0

summon_Elevated = 0

monster_attack = 10 * rng3
monster_HP = 15 * rng3
monster_def = 2 * rng3
monster_dead = 0

witch_HP = 0
witch_attack = 0
witch_defence = 0
witch_dead = 0

# introduction()

print("if you have a previous save type 'i' otherwise the save will be lost")
exit_potions = 1
true = 1
while true == 1:
    time.sleep(1)
    print("""
your options are
(s)earch forest
(v)iew stats
(e)nter store
(f)ight witch
(c)hallenge Elevated One""")
    command = input("")
    if command == "s":  # =========search forest=========
        rng1 = random.randint(0, 5)
        if rng1 == 1:
            rng1 = 0
            rng2 = random.randint(1, 4)
            print()
            print("you searched the forest and found", rng2, "candy(s)")
            candy += rng2
            rng2 = 0

        if rng1 != 1 and rng1 > 0:
            rng1 = 0
            rng3 = random.randint(2, 10)
            print()
            print("you stumbled across a level", rng3, "monster")
            command = input("""
your options are:
(f)ight monster
(r)un from monster
""")
            if command == "f":
                monster_attack = 10 * rng3
                monster_HP = 15 * rng3
                monster_def = 2 * rng3
                run = 0
                while run == 0:
                    if attack - monster_def < 0:
                        print("you dealt 0 damage, ", monster_HP, "monsters HP remains")
                        print()
                        HP = HP - (monster_attack - defence)
                        if monster_attack - defence < 0:
                            print("the monster dealt 0 damage, ", HP, "HP remains")
                            command = input("")
                            save_data()
                        if monster_attack - defence >= 0:
                            print("the monster dealt ", monster_attack - defence, "damage, ", HP, "HP remains")
                            input("")
                        if HP <= 0:
                            print("you died")
                            HP = 100
                            HP_max = 100
                            attack = 10
                            defence = 9
                            regen_pot_count = 0
                            candy = 0
                            break
                        print(" ")
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        print("(r)un (lose half your candies)")
                        command = input("")
                        save_data()
                    if attack - monster_def >= 0:
                        monster_HP = monster_HP - (attack - monster_def)
                        print()
                        print("you dealt", attack - monster_def, "damage, ", monster_HP, "monsters HP remains")
                        if monster_HP <= 0:
                            print("you killed the monster")
                            rng5 = random.randint(1, rng3)
                            print("it dropped", rng5, "candy")
                            candy += rng5
                            run = 1
                            monster_dead = 1
                            command = "~"
                        if monster_dead == 0:
                            if monster_attack - defence < 0:
                                print("the monster dealt 0 damage, ", HP, "HP remains")
                                print()
                                print("your options are:")
                                print("continue fight (any key)")
                                print("(u)se regen potion, ", regen_pot_count, "remaining")
                                print("(r)un (lose half your candies)")
                                command = input("")
                                save_data()
                            if monster_attack - defence >= 0:
                                HP = HP - (monster_attack - defence)
                                print("the monster dealt ", monster_attack - defence, "damage, ", HP, "HP remains")
                                if HP <= 0:
                                    print("you died")
                                    HP = 100
                                    HP_max = 100
                                    attack = 10
                                    defence = 9
                                    regen_pot_count = 0
                                    candy = 0
                                    break
                                print()
                                print("your options are:")
                                print("continue fight (any key)")
                                print("(u)se regen potion, ", regen_pot_count, "remaining")
                                print("(r)un (lose half your candies)")
                                command = input("")
                                save_data()
                                exit_potions = 1
                            if HP <= 0:
                                print("you died")
                                HP = 100
                                HP_max = 100
                                attack = 10
                                defence = 9
                                regen_pot_count = 0
                                candy = 0
                                break
                            if monster_HP <= 0:
                                print("you killed the monster")
                                rng5 = random.randint(1, rng3 + 1)
                                print("it dropped", rng5, "candy")
                                candy += rng5
                                run = 1
                        print()
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        print("(r)un (lose half your candies)")
                        command = input("")
                        save_data()
                    if command != "u" and command != "r" and command != "~":
                        print("")
                        print("you continue the fight with the monster")
                        exit_potions = 1
                    while command == "u" or exit_potions == 0:
                        exit_potions = 0
                        print("how many would you like to use")
                        try:
                            command = int(input(""))
                            save_data()
                            if command < regen_pot_count:
                                print("you restored ", 100 * command, "HP")
                                regen_pot_count += command - 2 * command
                                HP = HP + 100 * command
                                if HP > HP_max:
                                    HP = HP_max
                                    exit_potions = 1
                            if command > regen_pot_count and command != 0:
                                print()
                                print("you do not have enough ")
                                print()
                                exit_potions = 0
                            if command == 0:
                                exit_potions = 1
                        except ValueError:
                            print("invalid number")
                    if command == "r":
                        run = 1
                        command = ""
                        exit_potions = 1
                        if candy % 2 == 0:
                            candy = candy / 2
                            print("you ran from the monster")
                            print("you now have ", candy, " candies remaining")
                        if candy % 2 == 1:
                            candy = candy - 1
                            candy = candy / 2
                            print("you ran from the monster")
                            print("you now have ", candy, " candies remaining")
            if command == "r":
                print()
                print("you ran from the monster ")
    if command == "v":  # =========view stats=========
        print("your stats are:")
        print("HP = ", HP)
        print("max HP = ", HP_max)
        print("attack = ", attack)
        print("defence = ", defence)
        print("regen = ", regen_pot_count)
        print("candies = ", candy)

    leave_store = ""
    while command == "e" or leave_store == 1:  # =========enter store=========
        leave_store = 1
        print("you have ", candy, "candies")
        print("upgrade (a)ttack   (5 candy)")
        print("upgrade (d)efence   (5 candy)")
        print("upgrade (h)ealth   (5 candy)")
        print("buy (r)egen   (3 candy)")
        print("(l)eave store")
        command = input("")
        save_data()
        if command == "a" and candy >= 5:
            print("you upgraded your attack")
            attack += 5
            candy += -5
            print("you attack is now ", attack)
            print("you now have", candy, "candies remaining")
            print(" ")
            command = ""
        if command == "a" and candy < 5:
            print("you do not have enough candies")
            print(" ")
        if command == "d" and candy >= 5:
            print("you upgraded your defence")
            defence += 5
            candy += -5
            print("you defence is now ", defence)
            print("you now have", candy, "candies remaining")
            print(" ")
            command = ""
        if command == "d" and candy < 5:
            print("you do not have enough candies")
            print(" ")
        if command == "h" and candy >= 5:
            print("you upgraded your health")
            HP += 50
            HP_max += 50
            candy += -5
            print("your health is now ", HP)
            print("your max HP is now ", HP_max)
            print("you now have", candy, "candies remaining")
            print(" ")
            command = ""
        if command == "h" and candy < 5:
            print("you do not have enough candies")
            print(" ")
        if command == "r" and candy >= 3:
            print("you bought regen")
            regen_pot_count += 1
            candy += -3
            print("you regen count is now ", regen_pot_count)
            print("you now have", candy, "candies remaining")
            print(" ")
            command = ""
        if command == "r" and candy < 3:
            print("you do not have enough candies")
            print(" ")
        if command == "l":
            leave_store = 0
            print("you left the store")
            print(" ")
    if command == "f":  # =========fight witch=========
        print("which witch do you want to fight")
        print("(B)aba-Yaga ")
        print("(U)rsula ")
        print("(G)linda ")
        command = input("")
        save_data()
        if command == "B" or command == "b":  # =========Baba-Yaga=========
            witch_dead = 0
            witch_HP = 400
            witch_attack = 450
            witch_defence = 40
            run = 0
            witch_dead = 0
            while run == 0:
                exit_potions = 1
                if attack - witch_defence < 0:
                    print("you dealt 0 damage, ", witch_HP, "witch HP remains")
                    print()
                    if witch_attack - defence < 0:
                        print("the witch dealt 0 damage, ", HP, "HP remains")
                        print("")
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        command = input("")
                        save_data()
                    if witch_attack - defence >= 0:
                        print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                        HP = HP - (witch_attack - defence)
                        if HP <= 0:
                            print("you died")
                            HP = 100
                            HP_max = 100
                            attack = 10
                            defence = 9
                            regen_pot_count = 0
                            candy = 0
                            break
                    print()
                    print("your options are:")
                    print("continue fight (any key)")
                    print("(u)se regen potion, ", regen_pot_count, "remaining")
                    command = input("")
                    save_data()
                if attack - witch_defence >= 0:
                    print()
                    witch_HP = witch_HP - (attack - witch_defence)
                    print("you dealt", attack - witch_defence, "damage, ", witch_HP, "witch HP remains")
                    if witch_HP <= 0:
                        print("you killed the witch")
                        run = 1
                        command = ""
                        witch_dead = 1
                        if witch_B_dead == 0:
                            witch_B_dead = 1
                            print("it dropped an odd token")
                            print()
                    if witch_dead == 0:
                        if witch_attack - defence < 0:
                            print("the witch dealt 0 damage, ", HP, "HP remains")
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                        if witch_attack - defence >= 0:
                            HP = HP - (witch_attack - defence)
                            print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                            if HP <= 0:
                                print("you died")
                                HP = 100
                                HP_max = 100
                                attack = 10
                                defence = 9
                                regen_pot_count = 0
                                candy = 0
                                break
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                if command != "u":
                    print()
                    print("you continue the fight with the witch")
                    exit_potions = 1
                while command == "u" or exit_potions == 0:
                    exit_potions = 0
                    print("how many would you like to use")
                    try:
                        command = int(input(""))
                        save_data()
                        if command < regen_pot_count:
                            print("you restored ", 100 * command, "HP")
                            regen_pot_count += command - 2 * command
                            HP = HP + 100 * command
                            if HP > HP_max:
                                HP = HP_max
                                exit_potions = 1
                        if command > regen_pot_count and command != 0:
                            print()
                            print("you do not have enough ")
                            print()
                            exit_potions = 0
                        if command == 0:
                            exit_potions = 1
                    except ValueError:
                        print("invalid number")
        if command == "U" or command == "u":  # =========Ursula=========
            witch_HP = 200
            witch_attack = 600
            witch_defence = 60
            run = 0
            witch_dead = 0
            while run == 0:
                exit_potions = 1
                if attack - witch_defence < 0:
                    print("you dealt 0 damage, ", witch_HP, "witch HP remains")
                    print()
                    if witch_attack - defence < 0:
                        print("the witch dealt 0 damage, ", HP, "HP remains")
                        print("")
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        command = input("")
                        save_data()
                    if witch_attack - defence >= 0:
                        print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                        HP = HP - (witch_attack - defence)
                        if HP <= 0:
                            print("you died")
                            HP = 100
                            HP_max = 100
                            attack = 10
                            defence = 9
                            regen_pot_count = 0
                            candy = 0
                            break
                    print()
                    print("your options are:")
                    print("continue fight (any key)")
                    print("(u)se regen potion, ", regen_pot_count, "remaining")
                    command = input("")
                    save_data()
                if attack - witch_defence >= 0:
                    print()
                    witch_HP = witch_HP - (attack - witch_defence)
                    print("you dealt", attack - witch_defence, "damage, ", witch_HP, "witch HP remains")
                    if witch_HP <= 0:
                        print("you killed the witch")
                        run = 1
                        command = ""
                        witch_dead = 1
                        if witch_U_dead == 0:
                            witch_U_dead = 1
                            print("it dropped an odd token")
                            print()
                    if witch_dead == 0:
                        if witch_attack - defence < 0:
                            print("the witch dealt 0 damage, ", HP, "HP remains")
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                        if witch_attack - defence >= 0:
                            HP = HP - (witch_attack - defence)
                            print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                            if HP <= 0:
                                print("you died")
                                HP = 100
                                HP_max = 100
                                attack = 10
                                defence = 9
                                regen_pot_count = 0
                                candy = 0
                                break
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                if command != "u":
                    print()
                    print("you continue the fight with the witch")
                    exit_potions = 1
                while command == "u" or exit_potions == 0:
                    exit_potions = 0
                    print("how many would you like to use")
                    try:
                        command = int(input(""))
                        save_data()
                        if command < regen_pot_count:
                            print("you restored ", 100 * command, "HP")
                            regen_pot_count += command - 2 * command
                            HP = HP + 100 * command
                            if HP > HP_max:
                                HP = HP_max
                                exit_potions = 1
                        if command > regen_pot_count and command != 0:
                            print()
                            print("you do not have enough ")
                            print()
                            exit_potions = 0
                        if command == 0:
                            exit_potions = 1
                    except ValueError:
                        print("invalid number")
        if command == "G" or command == "g":  # =========Glinda=========
            witch_HP = 300
            witch_attack = 300
            witch_defence = 80
            run = 0
            witch_dead = 0
            while run == 0:
                exit_potions = 1
                if attack - witch_defence < 0:
                    print("you dealt 0 damage, ", witch_HP, "witch HP remains")
                    print()
                    if witch_attack - defence < 0:
                        print("the witch dealt 0 damage, ", HP, "HP remains")
                        print("")
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        command = input("")
                        save_data()
                    if witch_attack - defence >= 0:
                        print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                        HP = HP - (witch_attack - defence)
                        if HP <= 0:
                            print("you died")
                            HP = 100
                            HP_max = 100
                            attack = 10
                            defence = 9
                            regen_pot_count = 0
                            candy = 0
                            break
                    print()
                    print("your options are:")
                    print("continue fight (any key)")
                    print("(u)se regen potion, ", regen_pot_count, "remaining")
                    command = input("")
                    save_data()
                if attack - witch_defence >= 0:
                    print()
                    witch_HP = witch_HP - (attack - witch_defence)
                    print("you dealt", attack - witch_defence, "damage, ", witch_HP, "witch HP remains")
                    if witch_HP <= 0:
                        print("you killed the witch")
                        run = 1
                        command = ""
                        witch_dead = 1
                        if witch_G_dead == 0:
                            witch_G_dead = 1
                            print("it dropped an odd token")
                            print()
                    if witch_dead == 0:
                        if witch_attack - defence < 0:
                            print("the witch dealt 0 damage, ", HP, "HP remains")
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                        if witch_attack - defence >= 0:
                            HP = HP - (witch_attack - defence)
                            print("the witch dealt ", witch_attack - defence, "damage, ", HP, "HP remains")
                            if HP <= 0:
                                print("you died")
                                HP = 100
                                HP_max = 100
                                attack = 10
                                defence = 9
                                regen_pot_count = 0
                                candy = 0
                                break
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                if command != "u":
                    print()
                    print("you continue the fight with the witch")
                    exit_potions = 1
                while command == "u" or exit_potions == 0:
                    exit_potions = 0
                    print("how many would you like to use")
                    try:
                        command = int(input(""))
                        save_data()
                        if command < regen_pot_count:
                            print("you restored ", 100 * command, "HP")
                            regen_pot_count += command - 2 * command
                            HP = HP + 100 * command
                            if HP > HP_max:
                                HP = HP_max
                                exit_potions = 1
                        if command > regen_pot_count and command != 0:
                            print()
                            print("you do not have enough ")
                            print()
                            exit_potions = 0
                        if command == 0:
                            exit_potions = 1
                    except ValueError:
                        print("invalid number")
    if command == "c" and (witch_B_dead + witch_U_dead + witch_G_dead) != 3:  # =========challenge Elevated One=========
        print("the witches haven't been killed yet...")
    if command == "c" and (witch_B_dead + witch_U_dead + witch_G_dead) == 3:
        Elevated_life = 1
        while Elevated_life <= 6:
            if summon_Elevated == 0:
                summon_elevated_one()
                summon_Elevated = 1
            if Elevated_life == 1:
                elevated_HP = 1600
                elevated_attack = 1600
                elevated_defence = 320
                run = 0
                elevated_dead = 0
            if Elevated_life == 2:
                time.sleep(6)
                print('WOULDNT IT BE UNFAIR IF ONLY YOU COULD HEAL?')
                time.sleep(2)
                print("the Elevated One has recovered 800 HP")
                time.sleep(2)
                print("")
                elevated_HP = 800
                elevated_attack = 1600
                elevated_defence = 320
                run = 0
                elevated_dead = 0
            if Elevated_life == 3:
                time.sleep(6)
                print('"I CAN DO THIS ALL DAY"')
                time.sleep(2)
                print("the Elevated One has recovered 800 HP")
                time.sleep(2)
                print()
                elevated_HP = 800
                elevated_attack = 1600
                elevated_defence = 320
                run = 0
                elevated_dead = 0
            if Elevated_life == 4:
                time.sleep(6)
                print('"YOU FOOL YOU THINK YOU CAN DEFEAT ME?"')
                time.sleep(2)
                print("the Elevated One has recovered 400 HP")
                time.sleep(2)
                print()
                elevated_HP = 400
                elevated_attack = 1600
                elevated_defence = 320
                run = 0
                elevated_dead = 0
            if Elevated_life == 5:
                time.sleep(6)
                print('"DONT YOU EVER GIVE UP?"')
                time.sleep(2)
                print("the Elevated One has recovered 200 HP")
                time.sleep(2)
                print()
                elevated_HP = 200
                elevated_attack = 1600
                elevated_defence = 320
                run = 0
                elevated_dead = 0
            if Elevated_life == 6:
                time.sleep(6)
                print('"YOU HAVE KILLED ME ONCE BUT I WILL BE BACK"')
                time.sleep(2)
                run = 1
                Elevated_life = 7
                credits()
                true = 0
            while run == 0:
                exit_potions = 1
                if attack - elevated_defence < 0:
                    print("you dealt 0 damage, ", elevated_HP, "Elevated One HP remains")
                    print()
                    if elevated_attack - defence < 0:
                        print("the Elevated One dealt 0 damage, ", HP, "HP remains")
                        print("")
                        print("your options are:")
                        print("continue fight (any key)")
                        print("(u)se regen potion, ", regen_pot_count, "remaining")
                        command = input("")
                        save_data()
                    if elevated_attack - defence >= 0:
                        print("the Elevated One dealt ", elevated_attack - defence, "damage, ", HP, "HP remains")
                        HP = HP - (elevated_attack - defence)
                        if HP <= 0:
                            print("you died")
                            HP = 100
                            HP_max = 100
                            attack = 10
                            defence = 9
                            regen_pot_count = 0
                            candy = 0
                            break
                    print()
                    print("your options are:")
                    print("continue fight (any key)")
                    print("(u)se regen potion, ", regen_pot_count, "remaining")
                    command = input("")
                    save_data()
                if attack - elevated_defence >= 0:
                    print()
                    elevated_HP = elevated_HP - (attack - elevated_defence)
                    print("you dealt", attack - elevated_defence, "damage, ", elevated_HP, "Elevated One HP remains")
                    if elevated_HP <= 0:
                        print("the Elevated One has been killed")
                        run = 1
                        command = ""
                        elevated_dead = 1
                        Elevated_life = Elevated_life + 1
                    if elevated_dead == 0:
                        if elevated_attack - defence < 0:
                            print("the Elevated One dealt 0 damage, ", HP, "HP remains")
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                        if elevated_attack - defence >= 0:
                            HP = HP - (elevated_attack - defence)
                            print("the Elevated One dealt ", elevated_attack - defence, "damage, ", HP, "HP remains")
                            if HP <= 0:
                                print("you died")
                                HP = 100
                                HP_max = 100
                                attack = 10
                                defence = 9
                                regen_pot_count = 0
                                candy = 0
                                break
                            print("")
                            print("your options are:")
                            print("continue fight (any key)")
                            print("(u)se regen potion, ", regen_pot_count, "remaining")
                            command = input("")
                            save_data()
                if command != "u":
                    print()
                    exit_potions = 1
                while command == "u" or exit_potions == 0:
                    exit_potions = 0
                    print("how many would you like to use")
                    try:
                        command = int(input(""))
                        save_data()
                        if command < regen_pot_count:
                            print("you restored ", 100 * command, "HP")
                            regen_pot_count += command - 2 * command
                            HP = HP + 100 * command
                            if HP > HP_max:
                                HP = HP_max
                                exit_potions = 1
                        if command > regen_pot_count and command != 0:
                            print()
                            print("you do not have enough ")
                            print()
                            exit_potions = 0
                        if command == 0:
                            exit_potions = 1
                    except ValueError:
                        print("invalid number")

    if command == "i":
        try:
            files = "escape_forest_save.xlsx"
            wb = xl.load_workbook(files)
            sheet = wb['Sheet1']

            cell = sheet.cell(2, 2)
            HP = cell.value
            cell = sheet.cell(3, 2)
            HP_max = cell.value
            cell = sheet.cell(4, 2)
            attack = cell.value
            cell = sheet.cell(5, 2)
            defence = cell.value
            cell = sheet.cell(6, 2)
            regen_pot_count = cell.value
            cell = sheet.cell(7, 2)
            candy = cell.value
            cell = sheet.cell(8, 2)
            witch_B_dead = cell.value
            cell = sheet.cell(9, 2)
            witch_U_dead = cell.value
            cell = sheet.cell(10, 2)
            witch_G_dead = cell.value
            cell = sheet.cell(11, 2)
            summon_Elevated = cell.value
            print("previous save has been imported")
        except NameError:
            print("you do not have openpyxl module installed")
            print("you will not be able to import game data")
